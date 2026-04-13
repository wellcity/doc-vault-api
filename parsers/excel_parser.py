import sys
sys.path.append(".")

import uuid
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime, timezone
import json


def parse_xlsx(file_path: str, file_id: str, metadata: dict | None = None) -> tuple[list[dict], list[str]]:
    """
    解析 Excel (.xlsx) 檔案，回傳 (chunks, image_paths)
    每個工作表為一個 chunk，或每列為一個 chunk（視資料量決定）
    """
    chunks = []
    image_paths = []
    metadata = metadata or {}

    wb = load_workbook(file_path, data_only=True)

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]

        # 收集工作表內所有文字
        all_text = f"【工作表：{sheet_name}】\n"

        for row in ws.iter_rows(values_only=True):
            row_text = ""
            for cell in row:
                if cell is not None:
                    val = str(cell).strip()
                    if val:
                        row_text += val + " | "
            if row_text:
                all_text += row_text.rstrip(" | ") + "\n"

        if all_text.strip() == f"【工作表：{sheet_name}】":
            continue

        # 每工作表一個 chunk（也可改成每列）
        chunk = {
            "chunk_id": str(uuid.uuid4()),
            "file_id": file_id,
            "source_file": Path(file_path).name,
            "file_type": "excel",
            "page": sheet_idx,
            "chunk_index": 0,
            "text": all_text.strip(),
            "image_paths": [],
            "metadata_json": json.dumps({**metadata, "sheet_name": sheet_name}),
            "embedding": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        chunks.append(chunk)

    wb.close()
    return chunks, image_paths
