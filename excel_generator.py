"""
Excel 報表生成器
支援：多工作表、格式化標題、自動欄寬、交替列顏色、邊框
"""
import sys
sys.path.append(".")

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── 樣式工廠 ────────────────────────────────────────────────────────────────

def _make_border(style="thin"):
    side = Side(border_style=style, color="AAAAAA")
    return Border(left=side, right=side, top=side, bottom=side)


def _make_header_style(bg_color="1F497D"):
    return {
        "font": Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
        "fill": PatternFill("solid", fgColor=bg_color),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": _make_border(),
    }


def _make_data_style(idx, bg_even="FFFFFF", bg_odd="EEF3FA"):
    bg = bg_even if idx % 2 == 0 else bg_odd
    return {
        "font": Font(name="Calibri", size=10),
        "fill": PatternFill("solid", fgColor=bg),
        "alignment": Alignment(vertical="center"),
        "border": _make_border(),
    }


def _apply_style(cell, style):
    if "font" in style:
        cell.font = style["font"]
    if "fill" in style:
        cell.fill = style["fill"]
    if "alignment" in style:
        cell.alignment = style["alignment"]
    if "border" in style:
        cell.border = style["border"]


# ── 主函式 ──────────────────────────────────────────────────────────────────

def generate_excel_from_data(
    sheets: list[dict],
    output_name: str = "report",
    title: str | None = None,
) -> bytes:
    """
    根據資料結構生成 Excel 報表。

    Args:
        sheets: 工作表清單，格式如下：
            [{
                "name": "Sheet1",           # 工作表名稱
                "headers": ["欄位A", "欄位B", "欄位C"],  # 欄位標題
                "data": [
                    ["值1", "值2", "值3"],
                    ["值4", "值5", "值6"],
                ]
            }]

            進階：headers 可為 dict 設定格式
            ["欄位A", {"label": "欄位B", "width": 20, "align": "center"}]

            data 每列可為 list，或 dict（以 headers 為 key）
            data: [{"欄位A": "值1", "欄位B": "值2"}, ...]

        output_name: 輸出檔名（不含副檔名）
        title: 頂部標題列（選填），會自動跨欄

    Returns:
        Excel 檔案內容（bytes）
    """
    wb = Workbook()
    # 移除預設 sheet
    wb.remove(wb.active)

    # 全域配色
    PRIMARY = "1F497D"   # 深藍
    ACCENT = "2E86AB"    # 亮藍
    HEADER_BG = "1F497D"

    # ── 建立工作表 ──
    for sheet_cfg in sheets:
        ws = wb.create_sheet(title=sheet_cfg.get("name", "Sheet1"))

        headers_cfg = sheet_cfg.get("headers", [])
        data_rows = sheet_cfg.get("data", [])

        # 標題列（可選）
        if title:
            title_cell = ws.cell(row=1, column=1)
            col_count = len(headers_cfg) if headers_cfg else (data_rows[0] if data_rows else 1)
            title_cell.value = title
            title_cell.font = Font(name="Calibri", bold=True, size=14, color=PRIMARY)
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 28
            data_start_row = 2
        else:
            data_start_row = 1

        # 處理 headers 設定（str 或 dict）
        headers = []
        col_widths = []
        col_aligns = []

        for h in headers_cfg:
            if isinstance(h, dict):
                headers.append(h.get("label", str(h)))
                col_widths.append(h.get("width", 15))
                col_aligns.append(h.get("align", "left"))
            else:
                headers.append(str(h))
                col_widths.append(15)
                col_aligns.append("left")

        # 標題列
        header_style = _make_header_style(HEADER_BG)
        for col_idx, (h, width, align) in enumerate(
            zip(headers, col_widths, col_aligns), start=1
        ):
            cell = ws.cell(row=data_start_row, column=col_idx, value=h)
            _apply_style(cell, header_style)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[data_start_row].height = 20

        # 資料列
        for row_idx, row_data in enumerate(data_rows, start=1):
            row_num = data_start_row + row_idx

            # 資料格式：dict 或 list
            if isinstance(row_data, dict):
                row_values = [row_data.get(h, "") for h in headers]
            else:
                row_values = row_data

            row_style = _make_data_style(row_idx)

            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                _apply_style(cell, row_style)

                # 數字格式化
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

                # 對齊
                align = col_aligns[col_idx - 1] if col_idx <= len(col_aligns) else "left"
                cell.alignment = Alignment(
                    horizontal=align,
                    vertical="center",
                    wrap_text=True,
                )

            ws.row_dimensions[row_num].height = 18

    # 儲存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
